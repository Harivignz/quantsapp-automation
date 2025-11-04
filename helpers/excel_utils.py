import os
from openpyxl import load_workbook
import pandas as pd

def append_df_to_excel(filename, df, sheet_name="Data_Log", index=False):
    """
    Append DataFrame [df] into an existing Excel file [filename]
    into sheet [sheet_name]. If file doesn't exist, create it with headers once.
    Automatically avoids duplicated headers.
    """
    folder = os.path.dirname(filename)
    if folder:
        os.makedirs(folder, exist_ok=True)

    # If file does not exist, create it with headers
    if not os.path.exists(filename):
        df.to_excel(filename, sheet_name=sheet_name, index=index)
        return

    # If file exists, open and check if it has headers already
    book = load_workbook(filename)
    if sheet_name in book.sheetnames:
        ws = book[sheet_name]
        start_row = ws.max_row + 1
        # Append without headers
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=index,
                header=False,
                startrow=start_row
            )
    else:
        # Create sheet with headers if missing
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=index)
