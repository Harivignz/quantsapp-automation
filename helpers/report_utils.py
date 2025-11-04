# helpers/report_utils.py
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from helpers.notify import send_telegram


def update_summary(file_path: str):
    """
    Creates or updates a 'Daily_Summary' sheet in the Excel file
    showing per-day totals, avg, max, and min of Change %.
    Also sends a Telegram message with todayâ€™s summary.
    """
    if not os.path.exists(file_path):
        return

    try:
        # --- Read existing data log ---
        df = pd.read_excel(file_path, sheet_name="Data_Log")

        if df.empty or "Fetched At" not in df.columns:
            return

        # Ensure numeric conversion
        df["Change %"] = pd.to_numeric(df["Change %"], errors="coerce")

        # Extract date from "Fetched At"
        df["Date"] = pd.to_datetime(df["Fetched At"]).dt.date

        # --- Compute summary by date ---
        summary = (
            df.groupby("Date")["Change %"]
            .agg(["count", "mean", "max", "min"])
            .reset_index()
            .rename(
                columns={
                    "count": "Total Rows",
                    "mean": "Avg Change %",
                    "max": "Max Change %",
                    "min": "Min Change %",
                }
            )
        )

        # Round numeric columns
        summary["Avg Change %"] = summary["Avg Change %"].round(2)
        summary["Max Change %"] = summary["Max Change %"].round(2)
        summary["Min Change %"] = summary["Min Change %"].round(2)

        # --- Write back to Excel ---
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            summary.to_excel(writer, sheet_name="Daily_Summary", index=False)

        # --- Prepare Telegram message for today ---
        today = datetime.now().date()
        today_row = summary[summary["Date"] == today]
        if not today_row.empty:
            r = today_row.iloc[0]
            msg = (
                f"ðŸ“Š *Quantsapp Daily Summary ({today})*\n"
                f"â€¢ Total Entries: {int(r['Total Rows'])}\n"
                f"â€¢ Avg Change %: {r['Avg Change %']}\n"
                f"â€¢ Max Change %: {r['Max Change %']}\n"
                f"â€¢ Min Change %: {r['Min Change %']}\n"
                "âœ… Logged successfully in Excel."
            )
            send_telegram(msg)

    except Exception as e:
        send_telegram(f"âŒ Error updating daily summary: {e}")


def format_excel(file_path: str):
    """
    Auto-formats 'Data_Log' and 'Daily_Summary' sheets:
    - Bold headers
    - Adjust column widths
    - Green/Red colors for Change % columns
    """
    if not os.path.exists(file_path):
        return

    try:
        wb = load_workbook(file_path)

        # --- Format Data_Log ---
        if "Data_Log" in wb.sheetnames:
            ws = wb["Data_Log"]
            bold = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold
                cell.alignment = Alignment(horizontal="center")

            # Auto column width
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(val))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 35)

            # Colorize Change % columns
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if ws.cell(row=1, column=cell.column).value in ["Change %", "Price Change %"]:
                        try:
                            val = float(cell.value)
                            if val > 0:
                                cell.font = Font(color="008000")  # green
                            elif val < 0:
                                cell.font = Font(color="FF0000")  # red
                        except (ValueError, TypeError):
                            pass

        # --- Format Daily_Summary ---
        if "Daily_Summary" in wb.sheetnames:
            ws = wb["Daily_Summary"]
            bold = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(val))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

        wb.save(file_path)

    except Exception as e:
        send_telegram(f"âš ï¸ Excel formatting failed: {e}")
